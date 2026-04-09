# core/debug_export.py
"""
Excel debug matrix export — perfectly aligned with debug_render.py output.

Matching rules:
---------------------------------------------------------
FORBIDDEN               → black
✓                       → yellow
✓ (X)                   → yellow
X                       → recency colour
X (X) or "6"            → recency colour
SICK                    → light yellow
""                      → white (allowed unused)
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# Recency colours (1–12 months)
RECENCY_COLORS = [
    "A7D8FF", "73E0C1", "7CF37A", "C6F56F",
    "E4F55B", "FFD753", "FFB870", "FF8A8A",
    "FF5757", "AC6BFF", "9457D6", "6133A1"
]

BLACK = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def export_debug_matrix(path, people, areas, allowed_matrix, assignment, history_assignments):

    wb = Workbook()
    ws = wb.active
    ws.title = "DebugMatrix"

    # HEADER
    ws.cell(1, 1, "Area").font = Font(bold=True)
    for col, person in enumerate(people, start=2):
        c = ws.cell(1, col, person)
        c.font = Font(bold=True)
        ws.column_dimensions[c.column_letter].width = 18

    # reverse history so index 0 = last month
    rev_hist = list(reversed(history_assignments))

    # MAIN LOOP
    for row, area in enumerate(areas, start=2):
        ws.cell(row, 1, area).font = Font(bold=True)

        for col, person in enumerate(people, start=2):
            area_idx = row - 2
            person_idx = col - 2
            cell = ws.cell(row, col)

            # 1 — FORBIDDEN
            if allowed_matrix[area_idx][person_idx] == 0:
                cell.value = "FORBIDDEN"
                cell.fill = BLACK
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # Determine HISTORY AGE (recency)
            past_age = None
            for i, month in enumerate(rev_hist):
                if person_idx < len(month) and month[person_idx] == area_idx:
                    past_age = i + 1
                    break

            # 2 — CURRENT ASSIGNMENT
            if assignment[person_idx] == area_idx:

                # MATCH STREAMLIT: "✓" or "✓ (X)"
                if past_age is None:
                    cell.value = "✓"
                else:
                    cell.value = f"✓ ({past_age})"

                cell.fill = YELLOW
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # 3 — SICK
            if assignment[person_idx] is None:
                cell.value = "SICK"
                cell.fill = LIGHT_YELLOW
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # 4 — PAST ONLY (NOT CURRENT)
            if past_age is not None:
                idx = min(past_age - 1, len(RECENCY_COLORS) - 1)
                color = RECENCY_COLORS[idx]

                # MATCH STREAMLIT: age only → e.g., "6"
                cell.value = str(past_age)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # 5 — ALLOWED UNUSED
            cell.value = ""
            cell.fill = WHITE
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)