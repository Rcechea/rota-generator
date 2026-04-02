# core/io.py
"""
I/O utilities for loading and saving rota data.
Handles:
- Allowed matrix (CSV)
- Last month's assignments
- CSV and Excel export
- Fully compatible with Streamlit UploadedFile or normal file paths.
"""

import csv
import openpyxl
import io
from openpyxl.styles import Font, Alignment


# ------------------------------------------------------------
# Helper for cleaning CSV cells (removes BOM + quotes)
# ------------------------------------------------------------
def clean_cell(x: str):
    return (
        x.strip()
         .replace('"', '')
         .replace("\ufeff", '')   # UTF‑8 BOM
         .replace("ï»¿", '')      # UTF‑8 BOM Windows weirdness
    )


# ------------------------------------------------------------
# Utility: open file or UploadedFile
# ------------------------------------------------------------
def open_csv_source(path):
    """Return a file-like object for CSV reading."""
    # Streamlit UploadedFile has .read() / .getvalue()
    if hasattr(path, "read"):
        # Convert uploaded file bytes → string buffer
        return io.StringIO(path.getvalue().decode("utf-8-sig"))
    else:
        # Normal file path
        return open(path, newline="", encoding="utf-8-sig")


# ------------------------------------------------------------
# LOADING FUNCTIONS
# ------------------------------------------------------------
def load_allowed_matrix(path):
    """Load allowed-matrix CSV (0/1 grid)."""
    matrix = []
    f = open_csv_source(path)
    reader = csv.reader(f)

    for row in reader:
        clean_row = [int(clean_cell(x)) for x in row if clean_cell(x) != ""]
        matrix.append(clean_row)

    return matrix


def load_last_month(path):
    """Load last-month area indices, handling Streamlit uploads & BOM."""
    f = open_csv_source(path)
    raw = f.read().strip()

    if raw == "":
        return []

    parts = raw.replace("\n", ",").split(",")

    clean = []
    for x in parts:
        c = clean_cell(x)
        if c == "":
            clean.append(None)
        else:
            clean.append(int(c))

    return clean


# ------------------------------------------------------------
# EXPORT FUNCTIONS
# ------------------------------------------------------------
def export_csv(path, assignment, people, areas):
    """Export final assignment to CSV."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Person", "Assigned Area"])
        for person_idx, area_idx in enumerate(assignment):
            writer.writerow([people[person_idx], areas[area_idx]])


def export_excel(path, assignment, people, areas):
    """Export formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rota"

    ws["A1"] = "Person"
    ws["B1"] = "Assigned Area"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    for row_i, (person_idx, area_idx) in enumerate(enumerate(assignment), start=2):
        ws[f"A{row_i}"] = people[person_idx]
        ws[f"B{row_i}"] = areas[area_idx]

    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 25

    for row in ws.iter_rows(min_row=1, max_row=len(assignment) + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    wb.save(path)